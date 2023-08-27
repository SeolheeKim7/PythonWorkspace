from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")  # sample.xlsx file open
ws = wb.active  # active s

a1 = ws["A1"]  # A1 cell, number
b1 = ws["B1"]  # B1 cell, English
c1 = ws["C1"]  # C1 cell, Math

ws.column_dimensions["A"].width = 5  # A column width 5
ws.row_dimensions[1].height = 50  # 1 row height 50

# font color red, italic, bold
a1.font = Font(color="FF0000", italic=True, bold=True)
b1.font = Font(color="CC33FF", name="Arial", strike=True)  # strike line
c1.font = Font(color="0000FF", size=20, underline="single")  # underline

# border
thin_border = Border(left=Side(style="thin"),  # left thin
                     right=Side(style="thin"),  # right thin
                     top=Side(style="thin"),  # top thin
                     bottom=Side(style="thin"))  # bottom thin
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# fill the color
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.column == 1:  # A column
            continue
        # if 90 points or more, green
        if isinstance(cell.value, int) and cell.value >= 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
            cell.font = Font(color="FF0000")
        # # if 80 points or more, yellow
        # elif isinstance(cell.value, int) and cell.value >= 80:
        #     cell.fill = PatternFill(fgColor="FFFF00", fill_type="solid")
        # # if 70 points or more, blue
        # elif isinstance(cell.value, int) and cell.value >= 70:
        #     cell.fill = PatternFill(fgColor="0000FF", fill_type="solid")
        # # else red
        # else:
        #     cell.fill = PatternFill(fgColor="FF0000", fill_type="solid")

# freeze panes
ws.freeze_panes = "B2"  # freeze B2 cell
wb.save("sample_style.xlsx")  # save
