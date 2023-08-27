from openpyxl import Workbook
wb = Workbook()  # create new file
ws = wb.active  # ctive sheet

# row 1 data
# ws["A1"] = "Student Number"
# ws["B1"] = "Attendance"
# ws["C1"] = "Quiz1"
# ws["D1"] = "Quiz2"
# ws["E1"] = "Midterm"
# ws["F1"] = "Final"
# ws["G1"] = "Project"
# ws["H1"] = "Total"
# ws["I1"] = "Grade"
ws.append(["Student Number", "Attendance", "Quiz1", "Quiz2",
          "Midterm", "Final", "Project"])

# Define the data to be added
data = [
    [1, 10, 8, 5, 14, 26, 12],
    [2, 7, 3, 7, 15, 24, 18],
    [3, 9, 5, 8, 8, 12, 4],
    [4, 7, 8, 7, 17, 21, 18],
    [5, 7, 8, 7, 16, 25, 15],
    [6, 3, 5, 8, 8, 17, 0],
    [7, 4, 9, 10, 16, 27, 18],
    [8, 6, 6, 6, 15, 19, 17],
    [9, 10, 10, 9, 19, 30, 19],
    [10, 9, 8, 8, 20, 25, 20]
]
# Iterate through the data and write it to the specified cells
for row_index, row_data in enumerate(data):
    for col_index, cell_value in enumerate(row_data):
        ws.cell(row=row_index + 2, column=col_index + 1, value=cell_value)

# change value to 10 for all cells in quiz2 column
for col in ws.iter_cols(min_col=4, max_col=4, min_row=2):  # 1st row
    for cell in col:
        # print(cell.value)
        cell.value = 10

# add sum and grade columns
ws["H1"] = "Total"  # add 2 columns
ws["I1"] = "Grade"
index = 1
for col in ws.iter_cols(min_col=8, max_col=8, min_row=2):  # 1st row
    for cell in col:
        # print(cell.value)
        index = index + 1
        cell.value = f"=SUM(B{index}:G{index})"

# add grade column
for col in ws.iter_cols(min_col=9, max_col=9, min_row=2):  # 1st row
    for cell in col:
        index = index + 1
        cell.value = f"=IF(H{index}>=90, \"A\", IF(H{index}>=80, \"B\", IF(H{index}>=70, \"C\", IF(H{index}>=60, \"D\", \"F\"))))"


wb.save("sample_quiz.xlsx")  # save
