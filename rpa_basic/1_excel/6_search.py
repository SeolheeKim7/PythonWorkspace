from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")  # sample.xlsx file open
ws = wb.active  # active sheet

for row in ws.iter_rows(min_row=2):  # all rows except 1st row
    # print(row[2].value)
    if int(row[1].value) > 80:
        print(row[0].value, "student's English score is", row[1].value)

for row in ws.iter_rows(max_row=1):  # 1st row
    for cell in row:
        # print(cell.value)
        if cell.value == "English":
            cell.value = "Computer"


wb.save("sample_modified.xlsx")
