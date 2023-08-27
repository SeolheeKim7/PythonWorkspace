from openpyxl import Workbook
wb = Workbook()  # create new file
# ws = wb.active  # ctive sheet
ws = wb.create_sheet()  # create new sheet
ws.title = "MySheet"  # sheet name change
# RGB color code   # set the color of the sheet tab
ws.sheet_properties.tabColor = "ff99ff"

ws1 = wb.create_sheet("YourSheet")  # create new sheet
ws2 = wb.create_sheet("NewSheet", 2)  # create new sheet at index 2

new_ws = wb["NewSheet"]  # get the sheet
print(wb.sheetnames)  # get all the sheet names

# sheet copy
new_ws["A1"] = "Test"  # A1 cell에 Test 입력
target = wb.copy_worksheet(new_ws)  # new_ws sheet copy
target.title = "Copied Sheet"

wb.save("sample.xlsx")
