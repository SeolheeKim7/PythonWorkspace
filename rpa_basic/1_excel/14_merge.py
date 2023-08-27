from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# merge A1 ~ B2
ws.merge_cells("B2:D2")
ws["B2"].value = "Merged Cell"

wb.save("sample_merge.xlsx")  # save
