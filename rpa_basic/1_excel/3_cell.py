from random import *
from openpyxl import Workbook
wb = Workbook()  # create new file
ws = wb.active  # ctive sheet
ws.title = "ShSheet"  # sheet name change

# enter 1 in A1
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

# enter 4 in B1
ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])  # A1 cell information
print(ws["A1"].value)  # A1 cell value
print(ws["A10"].value)  # None (empty cell)

# row = 1, 2, 3, ...
# column = A(1), B(2), C(3), ...
print(ws.cell(row=1, column=1).value)  # ws["A1"].value
print(ws.cell(row=1, column=2).value)  # ws["B1"].value

c = ws.cell(column=3, row=1, value=10)  # ws["C1"].value = 10
print(c.value)  # ws["C1"].value

# enter random number
index = 1
for x in range(1, 11):  # 10 rows
    for y in range(1, 11):  # 10 columns
        # ws.cell(row=x, column=y, value=randint(0, 100))  # 0~100 random number
        ws.cell(row=x, column=y, value=index)  # 0~100 random number
        index += 1


wb.save("sample.xlsx")
