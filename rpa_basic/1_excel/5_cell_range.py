from openpyxl.utils.cell import coordinate_from_string
from openpyxl import Workbook
from random import *

wb = Workbook()  # create new file
ws = wb.active  # ctive sheet

ws.append(["Number", "English", "Math"])  # append row
for i in range(1, 11):  # 10 rows
    ws.append([i, randint(0, 100), randint(0, 100)])  # append row
wb.save("sample.xlsx")

col_B = ws["B"]  # English column
# print(col_B)
# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"]  # English, Math column
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1]  # 1st row
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6]  # 2nd row ~ 6th row
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()


# row_range = ws[2:ws.max_row]  # 2nd row ~ last row
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ")  # A/10, AZ/250
#         xy = coordinate_from_string(cell.coordinate)  # A/10, AZ/250
#         # print(xy, end=" ")
#         print(xy[0], end="")  # A
#         print(xy[1], end=" ")  # 10
#         # print(ws.cell(xy).value, end=" ")
#     print()


# total rows
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[1].value)

# total columns
# print(tuple(ws.columns))
# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows():  # all rows
#     print(row[1].value)

# for column in ws.iter_cols():  # all columns
#     print(column[0].value)

for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    print(row[0].value, row[1].value)  # English, Math
    # print(row)

for col in ws.iter_cols(min_row=2, max_row=11, min_col=2, max_col=3):
    print(col[0].value, col[1].value)  # English, Math
    # print(col)

wb.save("sample.xlsx")
