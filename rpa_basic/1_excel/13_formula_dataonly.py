from openpyxl import load_workbook

# 1. data_only=False: save formula and data
# data_only=True: save only data, not formula
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # print formula without data_only=True attribute
# for row in ws.values:
#     for cell in row:
#         print(cell)  # print only data, not formula
#         # print(cell.value)  # print only data, not formula (same as above)

# 2. data_only=True: save only data, not formula

wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# print formula with data_only=True attribute
# print(ws["A2"].value) didn't evaluates data -> None
# print(ws["A3"].value) didn't evaluates data -> None
for row in ws.values:
    for cell in row:
        print(cell)  # print only data, not formula
        # print(cell.value)  # print only data, not formula (same as above)
