from openpyxl import load_workbook  # load workbook
wb = load_workbook("sample.xlsx")  # sample.xlsx file open
ws = wb.active  # active sheet

# cell data
# for x in range(1, 11):  # 10 rows
#     for y in range(1, 11):  # 10 columns
#         print(ws.cell(row=x, column=y).value, end=" ")  # 1 2 3 4 ...
#     print()  # enter

# cell data 2
for x in range(1, ws.max_row + 1):  # 10 rows
    for y in range(1, ws.max_column + 1):  # 10 columns
        print(ws.cell(row=x, column=y).value, end=" ")  # 1 2 3 4 ...
    print()  # enter
