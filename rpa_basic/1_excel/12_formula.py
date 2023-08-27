import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today()  # today date
ws["A2"] = "=SUM(1, 2, 3)"  # SUM function
ws["A3"] = "=AVERAGE(1, 2, 3)"  # AVERAGE function

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"  # SUM function

wb.save("sample_formula.xlsx")  # save
