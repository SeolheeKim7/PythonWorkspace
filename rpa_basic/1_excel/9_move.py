from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")  # sample.xlsx file open
ws = wb.active  # active s

# ws.move_range("B1:C11", rows=0, cols=1)  # B1:C11 -> C1:D11
# ws["B1"].value = "Korean"  # B1 cell -> Koean

ws.move_range("C1:C11", rows=5, cols=-1)  # C1:C11 -> B6:B16

wb.save("sample_korean.xlsx")
